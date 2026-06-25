"""
Bolão Copa do Mundo 2026 — CÓPIA DE APOSTAS
=============================================
Lê as apostas do Google Forms e copia para a aba "apostas" da planilha
local em formato legível (ex: "México 2x1").

ATENÇÃO: este módulo está 80% funcional. Pontos pendentes de revisão:
  - TODO: validar formatação de apostas com placar parcialmente preenchido
  - TODO: tratar caso onde apostador apostou mas não informou placar

Uso
----
  python bolao_apostas.py
"""

import csv
import io
import json
import re
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# ══════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════

ORDEM_FORMS = [
    "México x África do Sul", "Coreia do Sul x República Tcheca", "Canadá x Bósnia",
    "Estados Unidos x Paraguai", "Catar x Suíça", "Brasil x Marrocos", "Haiti x Escócia",
    "Austrália x Turquia", "Alemanha x Curaçao", "Países Baixos (Holanda) x Japão",
    "Costa do Marfim x Equador", "Suécia x Tunísia", "Espanha x Cabo Verde",
    "Bélgica x Egito", "Arábia Saudita x Uruguai", "Irã x Nova Zelândia",
    "França x Senegal", "Iraque x Noruega", "Argentina x Argélia", "Áustria x Jordânia",
    "Portugal x RD Congo", "Inglaterra x Croácia", "Gana x Panamá", "Uzbequistão x Colômbia",
    "República Tcheca x África do Sul", "Suíça x Bósnia", "Canadá x Catar",
    "México x Coreia do Sul", "Estados Unidos x Austrália", "Escócia x Marrocos",
    "Brasil x Haiti", "Turquia x Paraguai", "Países Baixos (Holanda) x Suécia",
    "Alemanha x Costa do Marfim", "Equador x Curaçao", "Tunísia x Japão",
    "Espanha x Arábia Saudita", "Bélgica x Irã", "Uruguai x Cabo Verde",
    "Nova Zelândia x Egito", "França x Iraque", "Noruega x Senegal",
    "Argentina x Áustria", "Jordânia x Argélia", "Portugal x Uzbequistão",
    "Inglaterra x Gana", "Panamá x Croácia", "Colômbia x RD Congo",
    "República Tcheca x México", "África do Sul x Coreia do Sul", "Suíça x Canadá",
    "Bósnia x Catar", "Escócia x Brasil", "Marrocos x Haiti", "Turquia x Estados Unidos",
    "Paraguai x Austrália", "Equador x Alemanha", "Curaçao x Costa do Marfim",
    "Japão x Suécia", "Tunísia x Países Baixos (Holanda)", "Cabo Verde x Arábia Saudita",
    "Uruguai x Espanha", "Noruega x França", "Senegal x Iraque", "Jordânia x Argentina",
    "Argélia x Áustria", "Egito x Irã", "Nova Zelândia x Bélgica", "Colômbia x Portugal",
    "RD Congo x Uzbequistão", "Panamá x Inglaterra", "Croácia x Gana",
]

CONFIG_PATH = Path(__file__).parent / "config.json"

PRIMEIRA_LINHA_APOSTAS = 3

ANTECEDENCIA_PADRAO_MIN = 30
JOGO_SEM_ANTECEDENCIA   = 0

COL_INICIO_FORMS = 3


def carregar_config():
    if not CONFIG_PATH.exists():
        sys.exit(
            f"❌ Não encontrei {CONFIG_PATH.name}.\n"
            f"   Copie config.example.json para config.json e preencha "
            f"com os dados do seu bolão."
        )
    with open(CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    faltando = [k for k in ("id_forms", "planilha_local") if k not in cfg]
    if faltando:
        sys.exit(f"❌ Faltam chaves em config.json: {', '.join(faltando)}")
    cfg.setdefault("aba_forms", 0)
    return cfg


# ══════════════════════════════════════════════════════════
# NORMALIZAÇÃO E INTERPRETAÇÃO DE TEXTO LIVRE
# ══════════════════════════════════════════════════════════

def normalizar_texto(texto):
    """Remove acentos, espaços nas pontas e deixa em minúsculas."""
    if not texto:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


def parse_placar(texto):
    """Extrai (gols_a, gols_b) dos dois primeiros inteiros de uma string."""
    if not texto:
        return None, None
    nums = re.findall(r"\d+", str(texto))
    if len(nums) >= 2:
        return int(nums[0]), int(nums[1])
    return None, None


def determinar_vencedor_placar(gols_a, gols_b):
    """'casa', 'fora' ou 'empate' a partir de um placar."""
    if gols_a is None or gols_b is None:
        return None
    if gols_a > gols_b:
        return "casa"
    if gols_b > gols_a:
        return "fora"
    return "empate"


def time_mencionado(texto_norm, time_norm):
    """True se `time_norm` aparece em `texto_norm` respeitando fronteira de palavra."""
    if not time_norm:
        return False
    if time_norm in texto_norm:
        return True
    for token in time_norm.split():
        if len(token) > 2 and re.search(r"\b" + re.escape(token) + r"\b", texto_norm):
            return True
    return False


def extrair_vencedor_aposta(texto_vencedor, time_casa_norm, time_fora_norm):
    """Interpreta o campo 'em quem você aposta?' -> 'casa'/'fora'/'empate'/None."""
    if not texto_vencedor:
        return None
    t = normalizar_texto(texto_vencedor)
    if "empate" in t or "draw" in t or t in ("e", "x", "-"):
        return "empate"
    casa = time_mencionado(t, time_casa_norm)
    fora = time_mencionado(t, time_fora_norm)
    if casa and not fora:
        return "casa"
    if fora and not casa:
        return "fora"
    return None


def parse_timestamp(valor):
    """Converte o timestamp do Forms em datetime."""
    if not valor:
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(str(valor).strip(), fmt)
        except ValueError:
            continue
    return None


# ══════════════════════════════════════════════════════════
# LEITURA DAS FONTES DE DADOS
# ══════════════════════════════════════════════════════════

def ler_forms(id_planilha, indice_aba):
    """Baixa as respostas do Forms via exportação pública em CSV."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{id_planilha}"
        f"/gviz/tq?tqx=out:csv&gid={indice_aba}"
    )
    resposta = requests.get(url, timeout=30)
    resposta.raise_for_status()
    linhas = list(csv.reader(io.StringIO(resposta.text)))
    return linhas[0], linhas[1:]


def ler_jogos(caminho):
    """Lê jogos da aba 'jogos' da planilha local."""
    wb = load_workbook(caminho, data_only=True)
    ws_j = wb["jogos"]
    linhas_jogos = list(ws_j.iter_rows(min_row=2, values_only=True))

    jogos = []
    for i, nome in enumerate(ORDEM_FORMS):
        row = linhas_jogos[i] if i < len(linhas_jogos) else (None, None, None)
        data_jogo  = row[1] if len(row) > 1 else None
        placar_str = row[2] if len(row) > 2 else None
        gols_casa, gols_fora = parse_placar(placar_str)
        time_casa, _, time_fora = str(nome).partition(" x ")
        jogos.append({
            "idx": i,
            "nome": nome,
            "time_casa": time_casa.strip(),
            "time_fora": time_fora.strip(),
            "time_casa_norm": normalizar_texto(time_casa),
            "time_fora_norm": normalizar_texto(time_fora),
            "data": data_jogo,
            "gols_casa": gols_casa,
            "gols_fora": gols_fora,
            "vencedor": determinar_vencedor_placar(gols_casa, gols_fora),
        })
    wb.close()
    return jogos


def mapear_linhas(ws, primeira_linha_dados):
    """Nome do apostador (coluna A) -> número da linha."""
    linhas = {}
    for i, row in enumerate(
        ws.iter_rows(min_row=primeira_linha_dados, values_only=True),
        start=primeira_linha_dados,
    ):
        if row[0]:
            linhas[str(row[0]).strip()] = i
    return linhas


# ══════════════════════════════════════════════════════════
# INTERPRETAÇÃO DAS APOSTAS DO FORMS
# ══════════════════════════════════════════════════════════

def parse_apostas(linhas_forms, jogos):
    """Interpreta as respostas do Forms e retorna dict por apostador."""
    apostas = {}

    for linha in linhas_forms:
        if not linha or not any(linha):
            continue

        data_aposta = parse_timestamp(linha[0])
        nome = str(linha[1]).strip() if len(linha) > 1 and linha[1] else None
        if not nome:
            continue

        apostas[nome] = {"data_aposta": data_aposta, "jogos": {}}

        col = COL_INICIO_FORMS
        for jogo in jogos:
            venc_txt   = linha[col]     if col     < len(linha) else None
            placar_txt = linha[col + 1] if col + 1 < len(linha) else None
            col += 2

            if not venc_txt and not placar_txt:
                apostas[nome]["jogos"][jogo["idx"]] = None
                continue

            venc_ap = extrair_vencedor_aposta(
                venc_txt, jogo["time_casa_norm"], jogo["time_fora_norm"]
            )
            gc, gf = parse_placar(placar_txt)

            if venc_ap == "fora" and gc is not None:
                gc, gf = gf, gc

            venc_pelo_placar = determinar_vencedor_placar(gc, gf)
            if venc_ap and venc_pelo_placar and venc_ap != venc_pelo_placar:
                print(
                    f"  ⚠ Contradição em '{nome}' / '{jogo['nome']}': campo de "
                    f"vencedor indica '{venc_ap}', mas o placar {gc}x{gf} indica "
                    f"'{venc_pelo_placar}'. Mantendo o campo de vencedor."
                )

            apostas[nome]["jogos"][jogo["idx"]] = {
                "vencedor": venc_ap,
                "gols_casa": gc,
                "gols_fora": gf,
            }

    return apostas


# ══════════════════════════════════════════════════════════
# FORMATAÇÃO DA APOSTA EM TEXTO LEGÍVEL
# ══════════════════════════════════════════════════════════

def formatar_aposta(aposta, jogo):
    """
    Transforma a aposta estruturada em texto legível: 'Time 2x1' ou '-'.

    TODO: revisar casos onde gols_casa/gols_fora é None mas vencedor foi
    informado — atualmente exibe 'Time ?' o que pode confundir na planilha.
    """
    if not aposta:
        return "-"
    nome_time = {
        "casa":   jogo["time_casa"],
        "fora":   jogo["time_fora"],
        "empate": "Empate",
    }.get(aposta["vencedor"], "?")
    gc, gf = aposta["gols_casa"], aposta["gols_fora"]
    placar = f"{gc}x{gf}" if gc is not None else "?"
    return f"{nome_time} {placar}"


# ══════════════════════════════════════════════════════════
# ESCRITA NA PLANILHA LOCAL
# ══════════════════════════════════════════════════════════

BORDA = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def escrever_apostas(ws, apostas, jogos):
    """
    Aba 'apostas': escreve horário da aposta (col. B) e o texto de cada
    aposta (col. C em diante, uma por jogo).

    TODO: verificar alinhamento das colunas quando a aba 'apostas' tem
    jogos fora de ordem em relação a ORDEM_FORMS.
    """
    for nome, linha in mapear_linhas(ws, PRIMEIRA_LINHA_APOSTAS).items():
        dado = apostas.get(nome)

        # Coluna B: data/hora da aposta
        cel_data = ws.cell(row=linha, column=2)
        cel_data.font      = Font(name="Arial", size=9)
        cel_data.alignment = Alignment(horizontal="center", vertical="center")
        cel_data.border    = BORDA
        if dado and dado["data_aposta"]:
            cel_data.value         = dado["data_aposta"]
            cel_data.number_format = "DD/MM/YYYY HH:MM"
        else:
            cel_data.value = "-"

        # Colunas C+: uma por jogo
        for jogo in jogos:
            col = 3 + jogo["idx"]
            cel = ws.cell(row=linha, column=col)
            cel.font      = Font(name="Arial", size=9)
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border    = BORDA
            ap = dado["jogos"].get(jogo["idx"]) if dado else None
            cel.value = formatar_aposta(ap, jogo)


# ══════════════════════════════════════════════════════════
# EXECUÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════════

def main():
    cfg = carregar_config()

    print("=" * 60)
    print("  Bolão Copa do Mundo 2026 — CÓPIA DE APOSTAS")
    print("=" * 60)

    print("\n📋 Lendo jogos da planilha local...")
    jogos = ler_jogos(cfg["planilha_local"])
    print(f"   {len(jogos)} jogos cadastrados")

    print("\n🌐 Baixando respostas do Google Forms...")
    _, linhas_forms = ler_forms(cfg["id_forms"], cfg["aba_forms"])
    print(f"   {len(linhas_forms)} respostas recebidas")

    print("\n🔍 Interpretando apostas...")
    apostas = parse_apostas(linhas_forms, jogos)
    print(f"   {len(apostas)} apostadores processados")

    print("\n✍️  Atualizando aba 'apostas'...")
    wb = load_workbook(cfg["planilha_local"])
    escrever_apostas(wb["apostas"], apostas, jogos)

    cadastrados    = set(mapear_linhas(wb["apostas"], PRIMEIRA_LINHA_APOSTAS))
    nao_cadastrados = sorted(set(apostas) - cadastrados)
    if nao_cadastrados:
        print("\n⚠️  Responderam o Forms mas não estão na aba 'apostas':")
        for nome in nao_cadastrados:
            print(f"   - {nome}")

    wb.save(cfg["planilha_local"])
    print(f"\n✅ Planilha atualizada: {cfg['planilha_local']}")


if __name__ == "__main__":
    main()
