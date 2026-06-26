"""
Programa de gerenciamento de bolão da Copa.

Lê as abas "jogos" e "apostas" e preenche a aba "Apostadores" com:
  - "INVALIDO"  -> aposta feita fora do prazo permitido
  - 20          -> acertou o time/empate vencedor (10) E o placar exato (10)
  - 10          -> acertou só o vencedor
  - 0           -> errou o vencedor (jogo já com placar definido)
  - (vazio)     -> jogo ainda não tem placar definido (aguardando resultado)

Regras de prazo:
  - Para o jogo definido em JOGO_EXCECAO: aposta vale até a hora do jogo
    (inclusive). Depois da hora do jogo = INVALIDO.
  - Para todos os outros jogos: aposta vale até 30 minutos antes do jogo
    (inclusive). Menos de 30 min de antecedência = INVALIDO.

Regras de pontuação:
  - O texto da aposta é sempre "<Time ou Empate> <golsA>x<golsB>", onde
    golsA é o número de gols do time citado (ou do lado esquerdo, se for
    Empate) e golsB do adversário.
  - O "vencedor apostado" é sempre o nome do time (ou "Empate") citado no
    início do texto, independente do número do placar estar certo,
    incompleto ou mal digitado.
  - Os 10 pontos de placar exato só são dados se o número puder ser lido
    com confiança E coincidir exatamente com o placar real.

Cores aplicadas na aba Apostadores:
  - 20 pontos  -> fundo verde, texto branco negrito
  - 10 pontos  -> texto verde
  - 0 pontos   -> sem formatação especial
  - INVALIDO   -> fundo vermelho, texto branco negrito maiúsculo
"""
import re
import sys
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

JOGO_EXCECAO = 'México x África do Sul'
TOLERANCIA = timedelta(minutes=30)

RE_PLACAR = re.compile(r'^(\d+)\s*[xX]\s*(\d+)$')

VERDE_FUNDO   = PatternFill('solid', start_color='217346', end_color='217346')
VERMELHO_FUNDO = PatternFill('solid', start_color='C00000', end_color='C00000')
FONT_20       = Font(bold=True, color='FFFFFF')
FONT_10       = Font(color='217346')
FONT_INVALIDO = Font(bold=True, color='FFFFFF')


def aplicar_cor(cell, pontos_ou_str):
    """Aplica formatação conforme resultado."""
    if pontos_ou_str == 'INVALIDO':
        cell.font = FONT_INVALIDO
        cell.fill = VERMELHO_FUNDO
    elif pontos_ou_str == 20:
        cell.font = FONT_20
        cell.fill = VERDE_FUNDO
    elif pontos_ou_str == 10:
        cell.font = FONT_10


def carregar_jogos(ws_jogos):
    """Retorna dict: nome_do_jogo -> info (time1, time2, kickoff, vencedor, placar_par).
    
    Os nomes dos jogos na aba 'jogos' podem ser fórmulas de array que o openpyxl
    não calcula. Nesse caso, lemos os nomes diretamente dos cabeçalhos da aba
    'Apostadores' (linha 1, colunas C em diante), que contêm os nomes como texto.
    """
    jogos = {}
    for r in range(2, ws_jogos.max_row + 1):
        nome_cell = ws_jogos.cell(row=r, column=1).value
        kickoff = ws_jogos.cell(row=r, column=2).value
        placar_txt = ws_jogos.cell(row=r, column=3).value

        if not kickoff:
            continue

        # nome pode ser None se for ArrayFormula não calculada
        if nome_cell is None or hasattr(nome_cell, 'text'):
            nome = None
        else:
            nome = str(nome_cell).strip()

        vencedor = None
        placar_par = None
        if placar_txt:
            m = RE_PLACAR.match(str(placar_txt).strip())
            if m:
                a, b = int(m.group(1)), int(m.group(2))
                placar_par = (a, b)
                if a > b:
                    vencedor_idx = 'time1'
                elif b > a:
                    vencedor_idx = 'time2'
                else:
                    vencedor_idx = 'Empate'
            else:
                vencedor_idx = None
        else:
            vencedor_idx = None

        jogos[r - 1] = dict(
            nome=nome,
            kickoff=kickoff,
            vencedor_idx=vencedor_idx,
            placar_par=placar_par
        )
    return jogos


def carregar_jogos_com_nomes(ws_jogos, ws_apostadores):
    """Combina nomes dos cabeçalhos de Apostadores com dados de jogos."""
    # Nomes dos jogos estão nos cabeçalhos da aba Apostadores (linha 1, col 3+)
    nomes_por_col = {}
    for c in range(3, ws_apostadores.max_column + 1):
        v = ws_apostadores.cell(row=1, column=c).value
        if v:
            nomes_por_col[c - 2] = str(v).strip()  # índice 1-based desde o 1o jogo

    jogos_raw = carregar_jogos(ws_jogos)
    jogos = {}
    for idx, info in jogos_raw.items():
        nome = info.get('nome') or nomes_por_col.get(idx)
        if not nome:
            continue
        if ' x ' not in nome:
            continue
        time1, time2 = (p.strip() for p in nome.split(' x ', 1))

        vencedor = None
        if info['vencedor_idx'] == 'time1':
            vencedor = time1
        elif info['vencedor_idx'] == 'time2':
            vencedor = time2
        elif info['vencedor_idx'] == 'Empate':
            vencedor = 'Empate'

        jogos[nome] = dict(
            time1=time1,
            time2=time2,
            kickoff=info['kickoff'],
            vencedor=vencedor,
            placar_par=info['placar_par']
        )
    return jogos


def parse_aposta(texto, time1, time2):
    """Retorna (apostado, par_golos_ou_None)."""
    s = str(texto).strip()
    candidatos = sorted([time1, time2, 'Empate'], key=len, reverse=True)
    apostado = None
    for cand in candidatos:
        if s == cand or s.startswith(cand + ' '):
            apostado = cand
            break
    if apostado is None:
        return None, None

    resto = s[len(apostado):].strip()
    m = RE_PLACAR.match(resto)
    if not m:
        return apostado, None

    a, b = int(m.group(1)), int(m.group(2))
    if apostado == time2:
        par = (b, a)
    else:
        par = (a, b)
    return apostado, par


def processar(caminho_entrada, caminho_saida):
    wb_val = load_workbook(caminho_entrada, data_only=True)
    ws_jogos_v = wb_val['jogos']
    ws_apostas_v = wb_val['apostas']
    ws_apostadores_v = wb_val['Apostadores']

    jogos = carregar_jogos_com_nomes(ws_jogos_v, ws_apostadores_v)

    # Cabeçalhos da aba apostas: col 1 = apostador, col 2 = timestamp, col 3+ = jogos
    headers = {}
    for c in range(3, ws_apostas_v.max_column + 1):
        h = ws_apostas_v.cell(row=1, column=c).value
        if h:
            headers[c] = str(h).strip()

    wb_out = load_workbook(caminho_entrada, data_only=False)
    ws_apostadores = wb_out['Apostadores']
    ws_apostas_out = wb_out['apostas']

    stats = {'invalido': 0, '20': 0, '10': 0, '0': 0, 'pendente': 0, 'sem_jogo': 0}

    # Linhas de apostadores começam na linha 2 da aba apostas
    r = 2
    while r <= ws_apostas_v.max_row:
        nome = ws_apostas_v.cell(row=r, column=1).value
        ts = ws_apostas_v.cell(row=r, column=2).value

        # Pula linha se não tiver nome nem timestamp
        if not nome and not ts:
            r += 1
            continue

        for c, nome_jogo in headers.items():
            info = jogos.get(nome_jogo)
            if info is None:
                stats['sem_jogo'] += 1
                continue

            valor_aposta = ws_apostas_v.cell(row=r, column=c).value
            if valor_aposta is None or str(valor_aposta).strip() == '':
                continue

            kickoff = info['kickoff']
            if nome_jogo == JOGO_EXCECAO:
                invalido = ts > kickoff
            else:
                invalido = ts > (kickoff - TOLERANCIA)

            destino = ws_apostadores.cell(row=r, column=c)

            if invalido:
                destino.value = 'INVALIDO'
                aplicar_cor(destino, 'INVALIDO')
                stats['invalido'] += 1
                continue

            if info['vencedor'] is None:
                destino.value = None
                stats['pendente'] += 1
                continue

            apostado, par = parse_aposta(valor_aposta, info['time1'], info['time2'])
            pontos = 0
            if apostado == info['vencedor']:
                pontos += 10
                if par is not None and par == info['placar_par']:
                    pontos += 10

            destino.value = pontos
            aplicar_cor(destino, pontos)
            if pontos == 20:
                stats['20'] += 1
            elif pontos == 10:
                stats['10'] += 1
            else:
                stats['0'] += 1

        r += 1

    wb_out.save(caminho_saida)
    return stats


if __name__ == '__main__':
    entrada = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\maua\Downloads\Resultados_bolão.xlsx"
    saida = sys.argv[2] if len(sys.argv) > 2 else entrada
    stats = processar(entrada, saida)
    print(stats)
